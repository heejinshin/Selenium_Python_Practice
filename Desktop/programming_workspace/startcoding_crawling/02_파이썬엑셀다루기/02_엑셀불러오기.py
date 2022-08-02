import openpyxl

fpath = r'/Users/shinheejin/Desktop/programming_workspace/2/startcoding_crawling/02_파이썬엑셀다루기/참가자_data.xlsx'

# 엑셀 불러오기 
wb = openpyxl.load_workbook(fpath)

# 엑셀시트선택 

ws = wb['오징어게임'] 

# 3) 데이터 수정하기 
ws['A3'] = 456 
ws['B3'] = '성기훈' 

#4) 엑셀 저장하기 
wb.save(fpath)